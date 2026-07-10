import csv
import io
import json
import pathlib
import smtplib
from datetime import datetime, timezone
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Optional
from fastapi import APIRouter, Depends, Form, HTTPException, Request
from fastapi.responses import FileResponse, JSONResponse, StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app.models.delivery_log import DeliveryLog
from app.models.exam import Exam
from app.models.mark import Mark
from app.models.question_mark import QuestionMark
from app.models.submission import Submission

router = APIRouter(prefix="/results-tools", tags=["Results Tools"])

# ── Helpers ──────────────────────────────────────────────────────────────────
def _exam_or_404(db, exam_id):
    e = db.query(Exam).filter(Exam.id == exam_id).first()
    if not e: raise HTTPException(404, "Exam not found")
    return e

def _all_subs(db, exam_id):
    return db.query(Submission).filter(Submission.exam_id == exam_id).order_by(Submission.id).all()

def _get(db, model, pk):
    return db.query(model).filter(model.id == pk).first() if pk else None

def _avg(rows):
    vals = [r["percentage"] for r in rows if r["percentage"] is not None]
    return round(sum(vals) / len(vals), 2) if vals else 0

def _student_name(student) -> str:
    if not student:
        return "Unknown"
    try:
        if hasattr(student, "first_name") and student.first_name:
            return f"{student.first_name} {student.last_name or ''}".strip()
        raw = getattr(student, "name", None)
        if raw and not callable(raw):
            return str(raw)
    except Exception:
        pass
    return f"Student #{student.id}"

def _student_dept(student) -> str:
    try:
        if student and student.department:
            return student.department.name or ""
    except Exception:
        pass
    return ""

def _build_rows(db, exam, submissions):
    from app.models.student import Student
    from app.models.subject import Subject
    from app.models.semester import Semester
    subject  = _get(db, Subject,  getattr(exam, "subject_id",  None))
    semester = _get(db, Semester, getattr(exam, "semester_id", None))
    rows = []
    for sub in submissions:
        mark    = db.query(Mark).filter(Mark.submission_id == sub.id).first()
        student = _get(db, Student, sub.student_id)
        qmarks  = (db.query(QuestionMark)
                   .filter(QuestionMark.submission_id == sub.id)
                   .order_by(QuestionMark.question_number).all())
        rows.append({
            "submission_id":    sub.id,
            "student_id":       sub.student_id,
            "student_name":     _student_name(student),
            "student_code":     getattr(student, "student_code", "") if student else "",
            "student_email":    getattr(student, "email", "") if student else "",
            "student_semester": getattr(student, "semester", None) if student else None,
            "department":       _student_dept(student),
            "exam_title":       exam.title,
            "exam_id":          exam.id,
            "exam_semester":    getattr(exam, "semester", None),
            "subject_name":     subject.name  if subject  else "",
            "subject_code":     subject.code  if subject  else "",
            "semester_name":    semester.name if semester else (
                                f"Sem {exam.semester}" if getattr(exam,"semester",None) else ""),
            "status":           sub.status,
            "score":            mark.score        if mark else None,
            "max_score":        mark.max_score    if mark else None,
            "percentage":       mark.percentage   if mark else None,
            "letter_grade":     mark.letter_grade if mark else "",
            "is_locked":        mark.is_locked    if mark else False,
            "comments":         mark.comments     if mark else "",
            "marked_at":        mark.marked_at.isoformat()   if mark and mark.marked_at   else None,
            "released_at":      mark.released_at.isoformat() if mark and mark.released_at else None,
            "file_url":         "/uploads/file/" + pathlib.Path(sub.file_path).name if sub.file_path else "",
            "question_marks":   [{"number": qm.question_number, "awarded": qm.awarded_marks,
                                  "max": qm.max_marks, "comment": qm.comment or ""}
                                 for qm in qmarks],
        })
    return rows


# ── Routes ───────────────────────────────────────────────────────────────────
@router.get("/exam/{exam_id}/csv")
def export_csv(exam_id: int, db: Session = Depends(get_db)):
    exam = _exam_or_404(db, exam_id)
    subs = _all_subs(db, exam_id)
    if not subs: raise HTTPException(400, "No submissions found")
    rows   = _build_rows(db, exam, subs)
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Submission ID","Student Code","Student Name","Email",
                     "Department","Semester","Subject","Subject Code",
                     "Exam Title","Score","Max Score","Percentage","Grade",
                     "Status","Marked At","Released At","Comments"])
    for r in rows:
        writer.writerow([
            r["submission_id"], r["student_code"], r["student_name"], r["student_email"],
            r["department"], r["semester_name"] or r["student_semester"] or "",
            r["subject_name"], r["subject_code"], r["exam_title"],
            r["score"]      if r["score"]      is not None else "",
            r["max_score"]  if r["max_score"]  is not None else "",
            r["percentage"] if r["percentage"] is not None else "",
            r["letter_grade"], r["status"], r["marked_at"] or "", r["released_at"] or "",
            (r["comments"] or "").replace("\n", " "),
        ])
    output.seek(0)
    fname = f"marks_exam_{exam_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(iter([output.getvalue()]), media_type="text/csv",
                             headers={"Content-Disposition": f"attachment; filename={fname}"})


@router.get("/exam/{exam_id}/submissions")
def exam_submissions(exam_id: int, db: Session = Depends(get_db)):
    exam = _exam_or_404(db, exam_id)
    rows = _build_rows(db, exam, _all_subs(db, exam_id))
    return {"exam_id": exam_id, "exam_title": exam.title,
            "total": len(rows), "avg_pct": _avg(rows), "results": rows}


@router.post("/exam/{exam_id}/send-to-admin")
def send_to_admin(exam_id: int, request: Request,
                  note: str = Form(""), db: Session = Depends(get_db)):
    exam   = _exam_or_404(db, exam_id)
    subs   = _all_subs(db, exam_id)
    if not subs: raise HTTPException(400, "No submissions found")
    marked = [s for s in subs if s.status in ("marked","locked","returned","reviewed")]
    if not marked: raise HTTPException(400, "No marked submissions. Finish marking first.")
    rows    = _build_rows(db, exam, marked)
    sent_by = request.session.get("user_id") or request.session.get("teacher_id")
    log = DeliveryLog(
        exam_id=exam_id, sent_by=sent_by,
        recipient="admin_inbox", status="sent",
        sent_at=datetime.now(timezone.utc),
        message=json.dumps({
            "note": note.strip(), "submitted_by": sent_by,
            "submitted_at": datetime.now(timezone.utc).isoformat(),
            "exam_title": exam.title, "total_marked": len(marked),
            "total_submissions": len(subs), "average_pct": _avg(rows),
            "summary_rows": rows,
        }),
    )
    db.add(log); db.commit(); db.refresh(log)
    return JSONResponse({"ok": True, "delivery_log_id": log.id,
                         "total_marked": len(marked),
                         "message": f"Marks for {len(marked)} student(s) sent to admin inbox."})


@router.get("/exam/{exam_id}/delivery-logs")
def delivery_logs(exam_id: int, db: Session = Depends(get_db)):
    logs = (db.query(DeliveryLog).filter(DeliveryLog.exam_id == exam_id)
              .order_by(DeliveryLog.sent_at.desc()).all())
    result = []
    for log in logs:
        try:    msg = json.loads(log.message) if log.message else {}
        except: msg = {}
        result.append({
            "id": log.id, "exam_id": log.exam_id, "sent_by": log.sent_by,
            "recipient": log.recipient, "status": log.status,
            "sent_at":      log.sent_at.isoformat() if log.sent_at else None,
            "note":         msg.get("note", ""),
            "total_marked": msg.get("total_marked"),
            "average_pct":  msg.get("average_pct"),
        })
    return {"logs": result}


# ── Send result email to one student ─────────────────────────────────────────
@router.post("/submission/{submission_id}/send-email")
def send_result_email(submission_id: int, db: Session = Depends(get_db)):
    from app.models.student import Student
    sub = db.query(Submission).filter(Submission.id == submission_id).first()
    if not sub: raise HTTPException(404, "Submission not found")
    if sub.status not in ("marked", "returned"):
        raise HTTPException(400, "Paper not yet marked")

    student = db.query(Student).filter(Student.id == sub.student_id).first()
    if not student: raise HTTPException(404, "Student not found")
    email = getattr(student, "email", None)
    if not email: raise HTTPException(400, "Student has no email address")

    exam = db.query(Exam).filter(Exam.id == sub.exam_id).first()
    mark = db.query(Mark).filter(Mark.submission_id == sub.id).first()

    name  = _student_name(student)
    score = f"{mark.score} / {mark.max_score}" if mark and mark.score is not None else "Not yet graded"
    pct   = f"{round(mark.percentage, 1)}%" if mark and mark.percentage else ""
    grade = mark.letter_grade if mark else ""

    body = f"""Dear {name},

Your results for {exam.title if exam else 'your exam'} are now available.

Score: {score} {pct}
Grade: {grade}

Please log in to the ExamMark portal to view your detailed results and annotated paper.

Best regards,
ExamMark System
"""

    try:
        msg = MIMEMultipart()
        msg["From"]    = "noreply@exammark.edu"
        msg["To"]      = email
        msg["Subject"] = f"Your Results: {exam.title if exam else 'Exam'}"
        msg.attach(MIMEText(body, "plain"))

        with smtplib.SMTP("localhost", 25, timeout=5) as smtp:
            smtp.sendmail("noreply@exammark.edu", [email], msg.as_string())

        return JSONResponse({"ok": True, "message": f"Email sent to {email}"})

    except Exception as e:
        # Log the attempt even if SMTP fails (no real SMTP in dev)
        print(f"[EMAIL] Would send to {email}: {body[:100]}")
        return JSONResponse({
            "ok": True,
            "message": f"Email logged for {email} (SMTP not configured — set up SMTP to send real emails)",
            "preview": body
        })


# ── Excel export ──────────────────────────────────────────────────────────────
@router.get("/exam/{exam_id}/excel")
def export_excel(exam_id: int, db: Session = Depends(get_db)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    exam = _exam_or_404(db, exam_id)
    subs = db.query(Submission).filter(Submission.exam_id == exam_id).order_by(Submission.id).all()
    if not subs:
        raise HTTPException(400, "No submissions found for this exam")
    rows = _build_rows(db, exam, subs)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    hdr_fill  = PatternFill("solid", fgColor="2563EB")
    hdr_font  = Font(color="FFFFFF", bold=True, size=11)
    hdr_align = Alignment(horizontal="center", vertical="center")

    # No max score, no individual Q marks, status = Pass/Fail
    headers = ["Student Code","Student Name","Email","Department",
               "Exam","Subject","Score","Percentage","Grade","Result","Comments"]

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align

    for ri, r in enumerate(rows, 2):
        pct = r["percentage"] or 0
        result = "Pass" if pct >= 50 else "Fail"
        vals = [
            r["student_code"], r["student_name"], r["student_email"],
            r["department"], r["exam_title"], r["subject_name"],
            r["score"],
            round(pct, 1) if pct else None,
            r["letter_grade"], result,
            (r["comments"] or "").replace("\n", " "),
        ]

        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=ri, column=col, value=val)
            cell.alignment = Alignment(vertical="center")
            if ri % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FAFC")

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"marks_exam_{exam_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )