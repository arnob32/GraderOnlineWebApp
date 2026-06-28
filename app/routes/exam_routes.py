# app/routes/exam_routes.py
import json
import os
from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from app.dependencies import get_db
from app.services.exam_service import (
    resolve_subject_and_code,
    create_exam_record,
    link_subject,
    save_questions,
    save_question_images,
    generate_pdfs,
)
from app.services.excel_import_service import import_students_from_excel
from app.utils import pdf_generator

router    = APIRouter(prefix="/api/exams", tags=["Exams"])
templates = Jinja2Templates(directory="app/templates")

os.makedirs(pdf_generator.TEMP_IMAGE_DIR, exist_ok=True)
os.makedirs("generated_pdfs", exist_ok=True)


@router.get("/create-ui")
def exam_create_ui(
    request:          Request,
    subject_id:       str | None = None,
    subject_name:     str | None = None,
    subject_code:     str | None = None,
    subject_semester: str | None = None,
    db:               Session = Depends(get_db),
):
    from app.models.teacher import Teacher
    tid     = request.session.get("user_id")
    teacher = None
    if tid:
        try:
            teacher = db.query(Teacher).filter(Teacher.id == int(tid)).first()
        except Exception:
            pass

    sem_name = ""
    if subject_semester and subject_semester.strip().isdigit():
        from app.models.semester import Semester
        sem = db.query(Semester).filter(
            Semester.id == int(subject_semester)).first()
        if sem:
            sem_name = sem.name or (
                str(sem.year_start) + " - " + str(getattr(sem, "year_end", ""))
            ).strip(" -")

    if not teacher:
        from app.models.teacher import Teacher as _T
        teacher = db.query(_T).first()

    teacher_id = teacher.id if teacher else 0
    return templates.TemplateResponse("create_exam.html", {
        "request":               request,
        "teacher_id":            teacher_id,
        "teacher_department_id": teacher.department_id if teacher else "",
        "subject_semester":      sem_name,
        "subject_semester_id":   subject_semester or "0",
        "pre_subject_id":        subject_id        or "",
        "pre_subject_code":      subject_code      or "",
        "pre_subject_name":      subject_name      or "",
    })


@router.get("/download/{filename}")
def download_exam_pdf(filename: str):
    safe_name = os.path.basename(filename)
    pdf_path  = os.path.join("generated_pdfs", safe_name)
    if not os.path.exists(pdf_path):
        raise HTTPException(404, "PDF not found")
    return FileResponse(
        path=pdf_path,
        media_type="application/pdf",
        filename=safe_name,
        headers={"Content-Disposition": f'attachment; filename="{safe_name}"'},
    )


@router.post("/preview-excel")
async def preview_excel(
    file: UploadFile = File(...),
    db:   Session    = Depends(get_db),
):
    """Preview students in an uploaded Excel file without saving to DB."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx or .xls files are accepted")
    try:
        import openpyxl, re, io
        content = await file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active

        header_row = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row and any(str(c).strip().lower() in ("mtknr", "name")
                           for c in row if c):
                header_row = i
                break

        if not header_row:
            raise HTTPException(422, "Cannot find header row (Name, Mtknr)")

        headers = [str(c).strip().lower() if c else "" for c in
                   next(ws.iter_rows(min_row=header_row,
                                     max_row=header_row, values_only=True))]

        def col(name):
            for i, h in enumerate(headers):
                if name in h:
                    return i
            return None

        idx_name  = col("name")
        idx_mtknr = col("mtknr")
        idx_email = col("e-mail") or col("email") or col("mail")

        students = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or all(c is None for c in row):
                continue
            raw_name = str(row[idx_name]).strip()  if idx_name  is not None and row[idx_name]  else ""
            mtknr    = str(row[idx_mtknr]).strip() if idx_mtknr is not None and row[idx_mtknr] else ""
            email    = str(row[idx_email]).strip() if idx_email is not None and row[idx_email] else ""
            if not raw_name or not mtknr:
                continue
            student_code = re.sub(r"\D", "", mtknr)
            if not student_code:
                continue
            if "," in raw_name:
                last, first = raw_name.split(",", 1)
                name = f"{first.strip()} {last.strip()}"
            else:
                name = raw_name
            students.append({
                "student_code": student_code,
                "name":         name,
                "email":        email or f"{student_code}@student.university.de",
            })
        return {"students": students, "count": len(students)}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Failed to parse Excel: {e}")


@router.post("/create")
async def create_exam(
    request:              Request,
    title:                str        = Form(...),
    course_code:          str        = Form(""),
    subject:              str        = Form(""),
    description:          str        = Form(""),
    department_id:        int        = Form(0),
    semester:             int        = Form(0),
    total_marks:          int        = Form(...),
    teacher_id:           int        = Form(0),
    questions_json:       str        = Form(...),
    reference_boxes_json: str        = Form("[]"),
    cover_rules:          str        = Form(""),
    exam_date_time:       str        = Form(""),
    student_list_file:    UploadFile = File(None),
    db: Session = Depends(get_db),
):
    # Parse questions
    try:
        questions = json.loads(questions_json)
        if not isinstance(questions, list):
            raise ValueError
    except Exception:
        raise HTTPException(400, "Invalid questions format")

    try:
        reference_boxes = json.loads(reference_boxes_json)
        if not isinstance(reference_boxes, list):
            reference_boxes = []
    except Exception:
        reference_boxes = []

    # Parse date/time
    exam_date = exam_time = ""
    if exam_date_time and exam_date_time.strip():
        parts = [p.strip() for p in exam_date_time.split(",")]
        if len(parts) >= 1: exam_date = parts[0]
        if len(parts) >= 2: exam_time = parts[1]

    # Create exam record
    resolved_code = resolve_subject_and_code(db, subject, course_code)
    exam = create_exam_record(
        db, title, resolved_code, description,
        department_id, semester, total_marks,
        teacher_id, cover_rules, exam_date, exam_time,
    )
    subject_id = link_subject(db, exam, subject, resolved_code)
    save_questions(db, exam.id, questions)

    form                 = await request.form()
    question_image_paths = await save_question_images(db, form, exam.id, questions)

    # Get students: Excel file or fallback to enrollment
    import_result = None
    has_excel = (
        student_list_file is not None
        and getattr(student_list_file, "filename", None)
        and student_list_file.filename.strip() not in ("", "undefined")
    )
    print(f"[EXAM] has_excel={has_excel} filename={getattr(student_list_file,'filename',None)}")

    if has_excel:
        import io
        file_content = await student_list_file.read()
        print(f"[EXAM] Excel file size: {len(file_content)} bytes")

        import_result = import_students_from_excel(
            io.BytesIO(file_content),
            db,
            default_semester=semester if semester > 0 else 1,
            default_department_id=department_id if department_id > 0 else None,
        )
        print(f"[EXAM] Excel import: created={import_result['created']} "
              f"updated={import_result['updated']} "
              f"total={len(import_result['students'])} "
              f"errors={import_result['errors']}")

        if import_result["errors"] and not import_result["students"]:
            raise HTTPException(422,
                f"Excel import failed: {import_result['errors'][0]}")

        from app.models.student import Student
        student_codes = [s["student_code"] for s in import_result["students"]]
        db_students = db.query(Student).filter(
            Student.student_code.in_(student_codes)).all()
        print(f"[EXAM] Found {len(db_students)} students in DB")
        students_entries = [(s, 1) for s in db_students]
    else:
        from app.services.exam_service import get_eligible_students
        students_entries = get_eligible_students(db, subject_id, semester)
        print(f"[EXAM] Fallback: {len(students_entries)} eligible students")

    if not students_entries:
        raise HTTPException(404,
            "No students found. Upload a Teilnehmerliste Excel file or ensure "
            "students are enrolled in this subject.")

    # Generate PDFs
    generated = generate_pdfs(
        db, exam, students_entries, questions,
        question_image_paths, reference_boxes, subject_id,
    )
    if not generated:
        raise HTTPException(500, "PDF generation failed for all students")

    if not generated:
        raise HTTPException(500, "PDF generation failed for all students")

    # Return ALL generated PDFs so client can download each one
    all_pdf_urls = [f"/api/exams/download/{f}" for f in generated]

    response = {
        "success":        True,
        "exam_id":        exam.id,
        "title":          exam.title,
        "students_count": len(students_entries),
        "pdfs_generated": len(generated),
        "pdf_url":        all_pdf_urls[0],           # first for preview
        "pdf_urls":       all_pdf_urls,              # ALL for bulk download
        "filename":       generated[0],
        "filenames":      generated,
    }
    if import_result:
        response["excel_import"] = {
            "created": import_result["created"],
            "updated": import_result["updated"],
            "skipped": import_result["skipped"],
            "errors":  import_result["errors"],
        }
    return JSONResponse(response)




@router.get("/download-zip/{exam_id}")
def download_exam_zip_by_id(exam_id: int, db: Session = Depends(get_db)):
    """Download all generated PDFs for an exam as a single ZIP file."""
    import zipfile as _zip
    import io as _io
    from app.models.exam import Exam

    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        raise HTTPException(404, "Exam not found")

    # Files named: exam_{exam_id}_student_{student_id}.pdf
    import glob as _glob
    pdf_pattern = os.path.join("generated_pdfs", f"exam_{exam_id}_student_*.pdf")
    pdf_files   = _glob.glob(pdf_pattern)
    print(f"[ZIP] Found {len(pdf_files)} PDFs for exam {exam_id}: pattern={pdf_pattern}")

    if not pdf_files:
        raise HTTPException(404, "No PDFs found for this exam")

    # Create ZIP in memory
    buf = _io.BytesIO()
    with _zip.ZipFile(buf, 'w', _zip.ZIP_DEFLATED) as zf:
        for pdf_path in sorted(pdf_files):
            fname = os.path.basename(pdf_path)
            if os.path.exists(pdf_path):
                zf.write(pdf_path, fname)

    buf.seek(0)
    safe = "".join(c for c in (exam.title or f"exam_{exam_id}") if c.isalnum() or c in " -_").strip()
    zip_name = f"{safe}_all_papers.zip"

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{zip_name}"'}
    )

@router.get("/exams-api")
def exams_api(db: Session = Depends(get_db)):
    """List all exams."""
    from app.models.exam import Exam
    exams = db.query(Exam).order_by(Exam.id.desc()).all()
    return [
        {
            "id":    e.id,
            "title": e.title,
        }
        for e in exams
    ]