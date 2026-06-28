# app/services/excel_import_service.py
import io
import re


def import_students_from_excel(
    file_obj,
    db,
    default_semester: int = 1,
    default_department_id: int = None,
) -> dict:
    """
    Parse an Excel student list and upsert students into the DB.
    Accepts bytes or BytesIO object.
    Returns dict with created, updated, skipped, errors counts and student list.
    """
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

    from app.models.student import Student

    # Accept both bytes and BytesIO
    if isinstance(file_obj, (bytes, bytearray)):
        file_obj = io.BytesIO(file_obj)

    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    # Find header row containing Mtknr or Name
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if row and any(str(c).strip().lower() in ("mtknr", "name")
                       for c in row if c):
            header_row = i
            break

    if not header_row:
        return {
            "created": 0, "updated": 0, "skipped": 0, "total": 0,
            "students": [],
            "errors": ["Cannot find header row (Name, Mtknr columns required)"]
        }

    headers = [
        str(c).strip().lower() if c else ""
        for c in next(ws.iter_rows(
            min_row=header_row, max_row=header_row, values_only=True))
    ]

    def col(name):
        for i, h in enumerate(headers):
            if name in h:
                return i
        return None

    idx_name  = col("name")
    idx_mtknr = col("mtknr")
    idx_email = col("e-mail") or col("email") or col("mail")

    if idx_name is None or idx_mtknr is None:
        return {
            "created": 0, "updated": 0, "skipped": 0, "total": 0,
            "students": [],
            "errors": ["Excel must have 'Name' and 'Mtknr' columns"]
        }

    created = updated = skipped = 0
    errors  = []
    students = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or all(c is None for c in row):
            continue

        raw_name = str(row[idx_name]).strip()  if idx_name  is not None and row[idx_name]  else ""
        mtknr    = str(row[idx_mtknr]).strip() if idx_mtknr is not None and row[idx_mtknr] else ""
        email    = str(row[idx_email]).strip() if idx_email is not None and row[idx_email] else ""

        if not raw_name or not mtknr:
            skipped += 1
            continue

        student_code = re.sub(r"\D", "", mtknr)
        if not student_code:
            skipped += 1
            continue

        # Parse name — handle "Last, First" format
        if "," in raw_name:
            last, first = raw_name.split(",", 1)
            first_name = first.strip()
            last_name  = last.strip()
        else:
            parts      = raw_name.split()
            first_name = parts[0] if parts else raw_name
            last_name  = " ".join(parts[1:]) if len(parts) > 1 else ""

        if not email:
            email = f"{student_code}@student.university.de"

        try:
            existing = db.query(Student).filter(
                Student.student_code == student_code).first()

            if existing:
                existing.first_name = first_name
                existing.last_name  = last_name
                existing.email      = email
                if default_department_id:
                    existing.department_id = default_department_id
                if default_semester:
                    existing.semester = default_semester
                updated += 1
                students.append({
                    "id":           existing.id,
                    "student_code": existing.student_code,
                    "first_name":   existing.first_name,
                    "last_name":    existing.last_name,
                    "email":        existing.email,
                })
            else:
                student = Student(
                    first_name    = first_name,
                    last_name     = last_name,
                    email         = email,
                    student_code  = student_code,
                    password_hash = student_code,
                    department_id = default_department_id,
                    semester      = default_semester,
                )
                db.add(student)
                db.flush()
                created += 1
                students.append({
                    "id":           student.id,
                    "student_code": student.student_code,
                    "first_name":   student.first_name,
                    "last_name":    student.last_name,
                    "email":        student.email,
                })
        except Exception as e:
            errors.append(f"Row {student_code}: {e}")
            skipped += 1
            try:
                db.rollback()
            except Exception:
                pass

    try:
        db.commit()
    except Exception as e:
        errors.append(f"Commit error: {e}")
        db.rollback()

    return {
        "created":  created,
        "updated":  updated,
        "skipped":  skipped,
        "total":    created + updated,
        "students": students,
        "errors":   errors,
    }